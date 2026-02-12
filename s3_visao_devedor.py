"""
Devedor360 v2 - Step 3: Visao Consolidada por Devedor
Substitui s3-CriaVisaoDevedor.py.
Agrega processos por entidade (CNPJ raiz ou CPF) e calcula indicadores.
Modular para uso com frontend.
"""

import os
import json
import glob
import traceback
from datetime import datetime

import pandas as pd

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from config import Config
from utils import (
    normalizar_documento, obter_raiz_cnpj, identificar_tipo_documento,
    is_justica_trabalho, parse_date, formatar_data_iso,
)


# ============================================================================
# Indicadores por entidade
# ============================================================================

def calcular_entity_id(row) -> str:
    """
    Determina ID unico da entidade:
      - CNPJ → raiz (8 digitos)
      - CPF  → CPF completo (11 digitos)
    """
    raiz = str(row.get("CNPJ Raiz", "")).strip()
    doc = str(row.get("CNPJ Completo", "")).strip()
    if raiz and len(raiz) >= 8 and raiz != "NA":
        return raiz[:8]
    if doc and len(doc) == 14:
        return doc[:8]
    if doc and len(doc) == 11:
        return doc
    return doc or "DESCONHECIDO"


def aggregate_por_entidade(grupo: pd.DataFrame) -> dict:
    """
    Calcula todos os indicadores para um grupo de processos da mesma entidade.
    """
    n = len(grupo)

    # Contagens
    flag_ext = grupo.get("Flag Extinto", pd.Series(dtype=int))
    qtd_ativos = int((flag_ext == 0).sum()) if not flag_ext.empty else n
    qtd_extintos = int((flag_ext == 1).sum()) if not flag_ext.empty else 0

    # Execucoes fiscais
    classe_col = grupo.get("Classe", pd.Series(dtype=str)).fillna("")
    is_ef = classe_col.str.contains("1116", na=False)
    qtd_ef = int(is_ef.sum())

    # Trabalhistas
    cnj_col = grupo.get("Numero CNJ", pd.Series(dtype=str)).fillna("")
    qtd_trab = int(cnj_col.apply(is_justica_trabalho).sum())

    # Polo ativo (nao EF)
    flag_reu = grupo.get("Flag Reu", pd.Series(dtype=int))
    qtd_polo_ativo_nao_ef = int(((flag_reu == 1) & (~is_ef)).sum()) if not flag_reu.empty else 0

    # Valores
    val_acao = pd.to_numeric(grupo.get("Valor Acao"), errors="coerce")
    val_corr = pd.to_numeric(grupo.get("Valor Corrigido"), errors="coerce")

    total_val_acao = float(val_acao.sum()) if not val_acao.isna().all() else 0
    total_val_corr = float(val_corr.sum()) if not val_corr.isna().all() else 0

    # Saldo liquido = valor ativos - valor extintos (corrigido)
    if not flag_ext.empty and not val_corr.isna().all():
        val_ativos = float(val_corr[flag_ext == 0].sum())
        val_extintos = float(val_corr[flag_ext == 1].sum())
    else:
        val_ativos = total_val_corr
        val_extintos = 0
    saldo_liquido = val_ativos - val_extintos

    maior_valor = float(val_corr.max()) if not val_corr.isna().all() else 0

    # Datas
    data_mais_antiga = ""
    data_mais_recente = ""
    data_ult_mov = ""
    try:
        datas_aj = grupo.get("Data Ajuizamento", pd.Series(dtype=str)).dropna()
        datas_parsed = []
        for d in datas_aj:
            try:
                datas_parsed.append(parse_date(str(d)))
            except Exception:
                pass
        if datas_parsed:
            data_mais_antiga = min(datas_parsed).strftime("%d/%m/%Y")
            data_mais_recente = max(datas_parsed).strftime("%d/%m/%Y")
    except Exception:
        pass
    try:
        datas_mov = grupo.get("Data Ultimo Movimento", pd.Series(dtype=str)).dropna()
        movs_parsed = []
        for d in datas_mov:
            try:
                movs_parsed.append(parse_date(str(d)))
            except Exception:
                pass
        if movs_parsed:
            data_ult_mov = max(movs_parsed).strftime("%d/%m/%Y")
    except Exception:
        pass

    # Tribunais e classes unicos
    tribunais = sorted(set(
        str(t).strip() for t in grupo.get("Tribunal", []) if str(t).strip()))
    classes = sorted(set(
        str(c).strip() for c in grupo.get("Classe", []) if str(c).strip()))

    # Origens consolidadas
    all_origens = set()
    for ori in grupo.get("Origens", []):
        for o in str(ori).split(","):
            o = o.strip()
            if o:
                all_origens.add(o)

    # ID Individuo (lista de todos os IDs que contribuiram)
    ids_individuos = sorted(set(
        str(i).strip() for i in grupo.get("ID Individuo", []) if str(i).strip()))

    return {
        "Entity ID": grupo.name if hasattr(grupo, "name") else "",
        "IDs Individuos": ", ".join(ids_individuos),
        "Qtd Processos": n,
        "Qtd Ativos": qtd_ativos,
        "Qtd Extintos": qtd_extintos,
        "Qtd Exec Fiscal": qtd_ef,
        "Qtd Trabalhista": qtd_trab,
        "Qtd Polo Ativo Nao EF": qtd_polo_ativo_nao_ef,
        "Total Valor Acao": round(total_val_acao, 2),
        "Total Valor Corrigido": round(total_val_corr, 2),
        "Saldo Liquido": round(saldo_liquido, 2),
        "Maior Valor Individual": round(maior_valor, 2),
        "Data Mais Antiga": data_mais_antiga,
        "Data Mais Recente": data_mais_recente,
        "Ultima Atualizacao": data_ult_mov,
        "Tribunais": " | ".join(tribunais),
        "Classes": " | ".join(classes),
        "Origens": ", ".join(sorted(all_origens)),
    }


# ============================================================================
# Classe principal
# ============================================================================

class VisaoDevedor:
    """
    Gera visao consolidada por devedor.

    CLI:
        df = VisaoDevedor(Config.from_env()).executar()

    Frontend:
        vis = VisaoDevedor(cfg, df_processos=meu_df, progress_callback=cb)
        df = vis.executar()
    """

    def __init__(self, config: Config = None,
                 df_processos: pd.DataFrame = None,
                 progress_callback=None):
        self.cfg = config or Config.from_env()
        self.cb = progress_callback
        self._df_processos = df_processos

    # ------------------------------------------------------------------
    #  Execucao principal
    # ------------------------------------------------------------------

    def executar(self) -> pd.DataFrame:
        """
        Pipeline completo:
          1. Le DataFrame de processos (do S2 ou parametro)
          2. Cria Entity ID
          3. Agrega por Entity ID
          4. Junta com dados da planilha de entrada
          5. Calcula estatisticas de download
          6. Salva Excel formatado
        Retorna DataFrame de visao por devedor.
        """
        cfg = self.cfg

        # 1. Obter DataFrame de processos
        df = self._obter_df_processos()
        if df.empty:
            print("[S3] DataFrame vazio - nenhum processo para agregar.")
            return pd.DataFrame()

        self._emit("s3_inicio", {"total_processos": len(df)})

        # 2. Entity ID
        df["Entity ID"] = df.apply(calcular_entity_id, axis=1)

        # 3. Agregacao
        results = []
        grouped = df.groupby("Entity ID")
        total_groups = len(grouped)
        for i, (eid, grupo) in enumerate(grouped):
            try:
                agg = aggregate_por_entidade(grupo)
                agg["Entity ID"] = eid
                results.append(agg)
            except Exception as e:
                if cfg.debug:
                    print(f"[S3-ERRO] Entity {eid}: {e}")
            if self.cb and (i % 50 == 0 or i == total_groups - 1):
                self._emit("s3_progresso", {"i": i + 1, "total": total_groups})

        df_agg = pd.DataFrame(results)
        if df_agg.empty:
            return df_agg

        # 4. Join com planilha de entrada
        df_agg = self._join_input(df_agg)

        # 5. Estatisticas de download (paginas baixadas por individuo)
        df_agg = self._add_download_stats(df_agg)

        # 6. Ordena por Saldo Liquido (desc)
        df_agg = df_agg.sort_values("Saldo Liquido", ascending=False).reset_index(drop=True)

        # 7. Salva Excel
        out_name = f"visao_devedor_{datetime.now():%Y%m%d_%H%M}.xlsx"
        out_path = os.path.join(cfg.output_dir, out_name)
        os.makedirs(cfg.output_dir, exist_ok=True)
        self._salvar_excel(df_agg, out_path)

        print(f"[S3] Salvo: {out_path}  ({len(df_agg):,} entidades)")
        self._emit("s3_fim", {"arquivo": out_path, "entidades": len(df_agg)})
        return df_agg

    # ------------------------------------------------------------------
    #  Obter DataFrame de processos
    # ------------------------------------------------------------------

    def _obter_df_processos(self) -> pd.DataFrame:
        """Tenta usar df passado por parametro, senao le ultimo Excel do s2."""
        if self._df_processos is not None and not self._df_processos.empty:
            return self._df_processos.copy()

        # Procura ultimo arquivo saida_processos_* no output_dir
        pattern = os.path.join(self.cfg.output_dir, "saida_processos_consolidados_*.xlsx")
        files = sorted(glob.glob(pattern), reverse=True)
        if files:
            print(f"[S3] Lendo: {files[0]}")
            return pd.read_excel(files[0], dtype=str)

        # Fallback: procura no diretorio atual
        pattern2 = os.path.join(".", "saida_processos_*.xlsx")
        files2 = sorted(glob.glob(pattern2), reverse=True)
        if files2:
            print(f"[S3] Lendo (fallback): {files2[0]}")
            return pd.read_excel(files2[0], dtype=str)

        return pd.DataFrame()

    # ------------------------------------------------------------------
    #  Join com planilha de entrada
    # ------------------------------------------------------------------

    def _join_input(self, df_agg: pd.DataFrame) -> pd.DataFrame:
        """Adiciona nome do cliente/devedor da planilha de entrada."""
        cfg = self.cfg
        path = cfg.input_file
        if not os.path.isfile(path):
            alt = os.path.join(os.path.dirname(os.path.abspath(__file__)), path)
            if os.path.isfile(alt):
                path = alt
            else:
                return df_agg

        try:
            if path.lower().endswith(".xls"):
                df_in = pd.read_excel(path, dtype=str, engine="xlrd")
            else:
                df_in = pd.read_excel(path, dtype=str)
        except Exception:
            return df_agg

        # Detecta colunas
        doc_col = nome_col = None
        for c in df_in.columns:
            cl = c.lower().strip()
            if "documento" in cl or "nr_doc" in cl or "cpf" in cl or "cnpj" in cl:
                doc_col = c
            if "nome" in cl:
                nome_col = c
        if not doc_col or not nome_col:
            return df_agg

        # Mapeamento entity_id -> nome
        eid_nome = {}
        for _, row in df_in.iterrows():
            d = normalizar_documento(str(row.get(doc_col, "")))
            n = str(row.get(nome_col, "")).strip()
            if d and n:
                tipo = identificar_tipo_documento(d)
                if tipo == "CNPJ":
                    eid_nome[d[:8]] = n
                elif tipo == "CPF":
                    eid_nome[d] = n

        if "Nome Devedor" not in df_agg.columns:
            df_agg["Nome Devedor"] = df_agg["Entity ID"].map(
                lambda x: eid_nome.get(str(x).strip(), ""))
        return df_agg

    # ------------------------------------------------------------------
    #  Estatisticas de download
    # ------------------------------------------------------------------

    def _add_download_stats(self, df_agg: pd.DataFrame) -> pd.DataFrame:
        """Conta paginas e detalhes por entity ID."""
        cfg = self.cfg
        output_dir = cfg.output_dir
        if not os.path.isdir(output_dir):
            return df_agg

        stats_map = {}
        for name in os.listdir(output_dir):
            d = os.path.join(output_dir, name)
            if not os.path.isdir(d):
                continue
            meta_path = os.path.join(d, "metadata.json")
            if not os.path.isfile(meta_path):
                continue
            try:
                with open(meta_path, "r", encoding="utf-8") as f:
                    meta = json.load(f)
                doc = normalizar_documento(meta.get("documento", ""))
                tipo = meta.get("tipo_documento", "")
                eid = doc[:8] if tipo == "CNPJ" and len(doc) == 14 else doc

                # Conta paginas e detalhes
                n_pages = 0
                n_dets = 0
                for root, _, files in os.walk(d):
                    for fn in files:
                        if fn.startswith("page_") and fn.endswith(".json"):
                            n_pages += 1
                        elif fn.endswith(".json") and not fn.startswith("page_") \
                                and fn not in ("metadata.json", "processos_unicos.json"):
                            n_dets += 1

                if eid not in stats_map:
                    stats_map[eid] = {"paginas": 0, "detalhes": 0}
                stats_map[eid]["paginas"] += n_pages
                stats_map[eid]["detalhes"] += n_dets
            except Exception:
                pass

        df_agg["Paginas Baixadas"] = df_agg["Entity ID"].map(
            lambda x: stats_map.get(x, {}).get("paginas", 0))
        df_agg["Detalhes Baixados"] = df_agg["Entity ID"].map(
            lambda x: stats_map.get(x, {}).get("detalhes", 0))

        return df_agg

    # ------------------------------------------------------------------
    #  Salvar Excel formatado
    # ------------------------------------------------------------------

    def _salvar_excel(self, df: pd.DataFrame, path: str):
        """Salva com formatacao condicional (openpyxl)."""
        df.to_excel(path, index=False, engine="openpyxl")

        if not HAS_OPENPYXL:
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb.active

            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2F5496", end_color="2F5496",
                                       fill_type="solid")
            border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))

            # Header
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = border

            # Dados
            money_cols = {"Total Valor Acao", "Total Valor Corrigido",
                          "Saldo Liquido", "Maior Valor Individual"}
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True)

                    header_val = ws.cell(row=1, column=col_idx).value
                    if header_val in money_cols:
                        try:
                            cell.value = float(cell.value) if cell.value else 0
                            cell.number_format = '#,##0.00'
                        except (ValueError, TypeError):
                            pass

                    # Highlight saldo negativo
                    if header_val == "Saldo Liquido":
                        try:
                            v = float(cell.value) if cell.value else 0
                            if v < 0:
                                cell.font = Font(color="FF0000", bold=True)
                        except (ValueError, TypeError):
                            pass

            # Ajusta larguras
            for col_idx in range(1, ws.max_column + 1):
                header_val = str(ws.cell(row=1, column=col_idx).value or "")
                width = max(len(header_val) + 2, 12)
                if header_val in money_cols:
                    width = 18
                ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 40)

            # Congela header
            ws.freeze_panes = "A2"

            # Filtro automatico
            ws.auto_filter.ref = ws.dimensions

            wb.save(path)
        except Exception as e:
            if self.cfg.debug:
                print(f"[S3-WARN] Formatacao Excel: {e}")

    def _emit(self, evt, data):
        if self.cb:
            try:
                self.cb(evt, data)
            except Exception:
                pass


# ============================================================================
# API de conveniencia
# ============================================================================

def executar_visao_devedor(config: Config = None,
                            df_processos: pd.DataFrame = None,
                            progress_callback=None) -> pd.DataFrame:
    """Executa Step 3 completo. Retorna DataFrame de visao por devedor."""
    return VisaoDevedor(config, df_processos, progress_callback).executar()


# ============================================================================
# CLI
# ============================================================================

if __name__ == "__main__":
    cfg = Config.from_env()
    print(f"[S3] Gerando visao do devedor...")
    df = executar_visao_devedor(cfg)
    print(f"\n=== S3 FINALIZADO ===")
    print(f"  Entidades: {len(df):,}")
    if not df.empty:
        cols = ["Entity ID", "Nome Devedor", "Qtd Processos",
                "Saldo Liquido", "Qtd Exec Fiscal"]
        cols_disp = [c for c in cols if c in df.columns]
        if cols_disp:
            print(df[cols_disp].head(10).to_string(index=False))
